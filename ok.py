from openpyxl import Workbook

wb = Workbook()

# ---------------- Sheet 1: login ----------------
ws1 = wb.active
ws1.title = "login"
ws1.append(["Email", "Password"])
ws1.append(["user@exmaple.com", "Admin@123"])
ws1.append(["mohit@yopmail.com", "Tester@123"])

# ---------------- Sheet 2: questions ----------------
ws2 = wb.create_sheet("questions")

ws2.append([
    "Question Type","seq","page no","min","max","question","instruction",
    "option1","option2","option3","option4","option5","option6","option7",
    "option8","option9","option10","option11","option12","option13",
    "option14","option15","option16"
])

ws2.append(["Open ended",1,1,"","","What is your name?","Personal Detail"])
ws2.append(["Single Select",2,1,"","","Are you the head of the household?","Personal Detail","Yes","No"])
ws2.append(["Open Ended",3,1,"","","If not, what is the name of the head of the household?","Personal Detail"])
ws2.append(["Open ended",4,1,"","","How many people are residing in this house?","Personal Detail"])
ws2.append(["Single Select",5,1,"","","Do you have a ration card?","Personal Detail","Yes","No","Can't Say/Not sure"])
ws2.append(["Single Select",6,1,"","","Is this a Blue Card (BPL Ration Card)?","Personal Detail","Yes","No","Can't Say/Not sure"])
ws2.append(["Single Select",7,1,"","","Does your family own any land in Punjab?","Personal Detail","Yes","No","Don't want to disclose"])
ws2.append(["Dropdown",8,1,"","","What type of land/s does your family own?","Personal Detail",
            "None","Commercial","Industrial","Institutional","Residential","Agriculture",
            "Non-Agriculture (Fallow Land)","Other (pls specify)"])
ws2.append(["Open ended",9,1,"","","What is the size of the land","Personal Detail"])

ws2.append(["Open ended",10,2,"","","Name of the family member",""])
ws2.append(["Dropdown",11,2,"","","What is this member’s relationship with the Head of Household?","",
            "Head of the Household (HoH)","Spouse","Parent","Parents-in-law","Son","Daughter",
            "Brother","Sister","Son-in-law","Daughter-in-law","Grandchild",
            "Brother-in-law / Sister-in-law","Other Relative","Friend",
            "Other (please specify)","Domestic Help"])
ws2.append(["Single Select",12,2,"","","Is this member currently living in this household?","","Yes","No"])
ws2.append(["Open ended",13,2,"","","Birth date of the family member",""])
ws2.append(["Open ended",14,2,"","","Age of the family member (Automate if birth date given)",""])
ws2.append(["Single Select",15,2,"","","Gender of the family member","","Male","Female","Other"])
ws2.append(["Dropdown",16,2,"","","Marital Status of the family member","",
            "Single","Married","Widow","Divorced/separated","Other"])
ws2.append(["Open ended",17,2,"","","Contact number of this family member",""])
ws2.append(["Dropdown",18,2,"","","Are they a voter registered at this address?","",
            "Yes","Not registered for voting anywhere",
            "Registered, but not from this address","Can't Say"])
ws2.append(["Dropdown",19,2,"","","If not registered here, where are they registered as a voter?","",
            "Somewhere else in Punjab","Other State"])
ws2.append(["Dropdown",20,2,"","","Which State","",
            "Andaman and Nicobar Islands","Arunachal Pradesh","Assam","Bihar","Chhattisgarh",
            "Goa","Gujarat","Delhi","Haryana","Himachal Pradesh","Jharkhand","Karnataka",
            "Kerala","Madhya Pradesh","Maharashtra","Manipur","Meghalaya","Mizoram",
            "Nagaland","Odisha","Punjab","Rajasthan","Sikkim","Tamil Nadu","Telangana",
            "Tripura","Uttar Pradesh","Uttarakhand","West Bengal","Jammu and Kashmir",
            "Lakshadweep","Puducherry"])

ws2.append(["Open ended",21,2,"","","Voter ID","Add the number / Refuse / Can't find"])
ws2.append(["Open ended",22,2,"","","Aadhar number",""])

ws2.append(["Dropdown",23,2,"","","What is the educational qualification?","",
            "No formal education","Primary (1-5)","Middle (6-8)",
            "Matric (9-10)","Senior Secondary (11-12)",
            "Diploma","Graduate","Post Graduation and Above"])
ws2.append(["Single Select",24,2,"","","Is this family member still studying?","","Yes","No"])
ws2.append(["Single Select",25,2,"","","Is this family member enrolled?","","School","College","Not sure"])
ws2.append(["Single Select",26,2,"","","If in school, is it government or private?","","Government","Private","Can't say"])
ws2.append(["Single Select",27,2,"","","Has govt school improved since 2022?","","Yes","No","No response"])
ws2.append(["Single Select",28,2,"","","If in college, is it government or private?","","Government","Private"])

ws2.append(["Dropdown",29,2,"","","Is the family member employed?","",
            "Employed","Unemployed","Student","Retired","Can't say","Other"])
ws2.append(["Single Select",30,2,"","","If retired, were they a govt employee?","","Yes","No","Can't Say"])
ws2.append(["Dropdown",31,2,"","","Type of employment","",
            "Private salaried","Government permanent","Government contractual",
            "Self-employed private","Self-employed agriculture","Can't say","Other"])
ws2.append(["Dropdown",32,2,"","","If govt employee, which govt?","",
            "Central","Punjab","Other State","Chandigarh UT","Other"])
ws2.append(["Open ended",33,2,"","","Monthly income",""])

ws2.append(["Dropdown",34,2,"","","Religion","",
            "Hindu","Sikh","Muslim","Christian","Other","Can't say"])
ws2.append(["Dropdown",35,2,"","","Caste category","",
            "General","OBC","SC","ST","Other","Can't say"])
ws2.append(["Dropdown",36,2,"","","Sub-caste","",
            "Banias","Brahmins","Khatris","Rajputs","Jatt Sikhs",
            "Mazbi Sikhs","Ravidassia","Valmiki","Other"])

ws2.append(["Single Select",37,2,"","","Kya aap kisi Dera ko maante ho?","","Yes","No","Can't Say"])
ws2.append(["Dropdown",38,2,"","","Which Dera?","","All known Deras","Other"])
ws2.append(["Single Select",39,2,"","","Kya aap kisi Pastor ko maante ho?","","Yes","No","Can't Say"])
ws2.append(["Dropdown",40,2,"","","Which Pastor?","","All known Pastors","Other"])

ws2.append(["Single Select",41,2,"","","Does member suffer from drug problem?","","Yes","No","Can't Say"])
ws2.append(["Dropdown",42,2,"","","How long suffered?","",
            "0-2 years","2-5 years","5-10 years","More than 10 years"])
ws2.append(["Single Select",43,2,"","","Do they go to treatment centre?","","Yes","No","Can't Say"])
ws2.append(["Single Select",44,2,"","","Which centre?","","Government","Private","Can't Say"])
ws2.append(["Dropdown",45,2,"","","Last visit to treatment centre","",
            "Within 3 months","3-6 months","6 months-1 year",
            "1-5 years","More than 5 years"])
ws2.append(["Dropdown",46,2,"","","Why did not go?","",
            "Didn't want to","Centre too far","Other"])
ws2.append(["Dropdown",47,2,"","","Why did they stop going?","",
            "Didn't want to","Centre too far","Other"])

ws2.append(["Single Select",48,2,"","","Is drug abuse a problem in area?","","Yes","No","Can't Say"])
ws2.append(["Dropdown",49,2,"","","Distance to govt treatment centre","",
            "Less than 5 km","5-10 km","10-20 km","More than 20 km","Not sure"])

ws2.append(["Single Select",50,22,"","","Support Yudh Nashe De Virudh campaign?","","Yes","No","Can't Say"])
ws2.append(["Single Select",51,2,"","","Impact on drug problem?","","Yes","No","Can't Say"])
ws2.append(["Dropdown",52,2,"","","Impact of campaign","",
            "Positive","Negative","No impact","Can't Say"])
ws2.append(["Single Select",53,3,"","","Has any member of this family migrated internationally?","","Yes","No","Can't Say"])

ws2.append(["Single Select",54,4,"","","Has the household taken any pending loans?","","Yes","No","Can't Say"])
ws2.append(["Open ended",55,4,"","","How many pending loans does the household have?",""])
ws2.append(["Dropdown",56,4,"","","Main purpose of the loan","",
            "Agriculture","Business","Home construction/repair","Education",
            "Marriage","Medical treatment","Vehicle purchase","Other","Can't Say"])
ws2.append(["Dropdown",57,4,"","","Source of the loan","",
            "Bank","Cooperative society","Moneylender",
            "Friends/family","Other","Can't Say"])
ws2.append(["Open ended",58,4,"","","How much loan was taken?",""])
ws2.append(["Open ended",59,4,"","","Interest rate for the loan",""])
ws2.append(["Open ended",60,4,"","","How much loan has been paid back?",""])
ws2.append(["Single Select",61,4,"","","Difficulty in repaying loan?","","Yes","No","Can't Say"])

ws2.append(["Dropdown",62,5,"","","Where does household seek medical treatment?","",
            "Government hospital/clinic","Aam Aadmi Clinic",
            "Private doctor/hospital","Traditional healer",
            "Self-medication","Other","Can't Say"])
ws2.append(["Single Select",63,5,"","","Difficulty accessing healthcare due to money?","","Yes","No","Can't Say"])
ws2.append(["Dropdown",64,5,"","","Distance to nearest govt health facility","",
            "Less than 1 km","1-5 km","5-10 km","More than 10 km","Not sure"])
ws2.append(["Single Select",65,5,"","","Has any member visited Aam Aadmi Clinic?","","Yes","No","Can't Say"])
ws2.append(["Single Select",66,5,"","","Liked Aam Aadmi Clinic services?","","Yes","No","Can't Say"])
ws2.append(["Single Select",67,5,"","","Household has Ayushman insurance?","","Yes","No","Can't Say"])
ws2.append(["Single Select",68,5,"","","Household has private insurance?","","Yes","No","Can't Say"])

ws2.append(["Dropdown",69,6,"","","Law and order situation","",
            "Very good","Good","Bad","Very bad","Can't Say"])
ws2.append(["Dropdown",70,6,"","","Common crimes in area","",
            "Theft","Drug-related crimes","Violence/assault",
            "Domestic violence","Other","No crimes","Can't Say"])
ws2.append(["Single Select",71,6,"","","Any household member faced crime in last year?","","Yes","No","Can't Say"])
ws2.append(["Dropdown",72,6,"","","Nature of crime","",
            "Theft","Drug-related","Violence/assault",
            "Domestic violence","Other","Can't Say"])

ws2.append(["Single Select",73,7,"","","Faced caste-based discrimination?","","Yes","No","Can't Say"])
ws2.append(["Dropdown",74,7,"","","When was last discrimination?","",
            "Within last year","1-2 years ago","2-5 years ago",
            "More than 5 years ago","Can't Say"])
ws2.append(["Dropdown",75,7,"","","Nature of caste discrimination","",
            "At work/school","Accessing services","Social exclusion",
            "Verbal abuse","Physical violence","Other","Can't Say"])
ws2.append(["Single Select",76,7,"","","Caste barriers in accessing services?","","Yes","No","Can't Say"])

ws2.append(["Single Select",77,7,"","","Benefited from 300 unit free bijli?","","Yes","No","Can't Say"])
ws2.append(["Open ended",78,"","","","Electricity meter number",""])
ws2.append(["Open ended",79,"","","","Monthly savings from bijli scheme",""])
ws2.append(["Single Select",80,"","","","Happy with bijli scheme?","","Yes","No","Can't Say"])

ws2.append(["Single Select",81,"","","","Benefited from Sarkar Tuade Dwar?","","Yes","No","Can't Say"])
ws2.append(["Single Select",82,"","","","Happy with this scheme?","","Yes","No","Can't Say"])

ws2.append(["Single Select",83,"","","","Benefited from Mukhyamantri Teerth Yatra?","","Yes","No","Can't Say"])
ws2.append(["Single Select",84,"","","","Happy with this scheme?","","Yes","No","Can't Say"])

ws2.append(["Single Select",85,"","","","Benefited from free treatment/medicines?","","Yes","No","Can't Say"])
ws2.append(["Open ended",86,"","","","Monthly savings from health scheme",""])
ws2.append(["Single Select",87,"","","","Happy with this scheme?","","Yes","No","Can't Say"])

ws2.append(["Single Select",88,"","","","Benefited from govt pension?","","Yes","No","Can't Say"])
ws2.append(["Open ended",89,"","","","Monthly pension amount",""])
ws2.append(["Single Select",90,"","","","Happy with this scheme?","","Yes","No","Can't Say"])

ws2.append(["Single Select",91,"","","","Benefited from Free Bus Yatra for women?","","Yes","No","Can't Say"])
ws2.append(["Open ended",92,"","","","Monthly savings from bus scheme",""])
ws2.append(["Single Select",93,"","","","Happy with this scheme?","","Yes","No","Can't Say"])

ws2.append(["Single Select",94,"","","","Benefited from CM di Yogshala?","","Yes","No","Can't Say"])
ws2.append(["Single Select",95,"","","","Happy with this scheme?","","Yes","No","Can't Say"])
ws2.append(["Single Select",96,"","","","Interested in free yoga after explanation?","","Yes","No","Can't Say"])

ws2.append(["Open ended",97,"","","","Who is your MLA",""])
ws2.append(["Single Select",98,"","","","Happy with local MLA?","","Yes","No","Can't Say"])
ws2.append(["Single Select",99,"","","","Happy with local Sarpanch/Councillor?","","Yes","No","Can't Say"])

ws2.append(["Single Select",100,"","","","Do you receive 24 hour electricity?","","Yes","Almost","No"])
ws2.append(["Dropdown",101,"","","","Daily power cut hours","",
            "0","1-2","2-6","6-12","More than 12","Can't Say"])
ws2.append(["Single Select",102,"","","","Electricity improved since 2022?","","Yes","No","Can't Say"])
ws2.append(["Open ended",103,"","","","Number of meter connections",""])
ws2.append(["Open ended",104,"","","","Electricity meter number",""])

ws2.append(["Dropdown",105,"","","","Main source of drinking water","",
            "Piped water","Borewell","Handpump","Public tap",
            "Water tank","River/Lake","Bottled","Other","Can't Say"])
ws2.append(["Dropdown",106,"","","","Water supply frequency","",
            "24x7","Once a day","Once every two days","Irregular"])
ws2.append(["Dropdown",107,"","","","Toilet access","",
            "Private toilet","Community toilet","No toilet"])
ws2.append(["Dropdown",108,"","","","Garbage pickup frequency","",
            "Daily","Every 2-3 days","Weekly","Irregular","Never","Can't Say"])
ws2.append(["Dropdown",109,"","","","Cleanliness satisfaction","",
            "Very satisfied","Somewhat satisfied","Not satisfied"])

ws2.append(["Single Select",110,"","","","Owns 2 wheeler?","","Yes","No","Not Sure"])
ws2.append(["Single Select",111,"","","","Owns car?","","Yes","No","Not Sure"])

ws2.append(["Dropdown",112,"","","","Sources of news","",
            "TV","Newspapers","Social Media","Mobile Apps",
            "Radio","Govt sources","Other","Can't Say"])
ws2.append(["Dropdown",113,"","","","Newspapers read","",
            "Punjabi Jagran","Ajit","Hindustan Times","Dainik Bhaskar",
            "Punjab Kesari","Economic Times","Do not read","Other","Can't Say"])
ws2.append(["Dropdown",114,"","","","News channels watched","",
            "DD Punjabi","PTC","ABP","Aaj Tak","NDTV","India TV",
            "Republic","Zee Punjab","TV Punjab","Other","Can't Say"])
ws2.append(["Dropdown",115,"","","","Entertainment channels","",
            "Star Plus","Sony","Zee","Colors","Star Sports",
            "PTC Punjabi","Other","Can't Say"])
ws2.append(["Dropdown",116,"","","","Social media used most","",
            "None","Whatsapp","Facebook","YouTube",
            "Instagram","Twitter","Other","Can't Say"])

ws2.append(["Dropdown",117,"","","","Urgent infrastructure needs","",
            "Roads","Transport","Drug rehab","Electricity",
            "Street lights","Water","Healthcare","School",
            "Waste management","Sewerage","Internet","Parks","Other","Can't Say"])

ws2.append(["Dropdown",118,"","","","Top 3 works of Punjab govt","",
            "Improved schools","Free electricity","Employment",
            "Drug action","Healthcare","Roads","Transport",
            "Women free bus","Farmer support",
            "Corruption free governance","Other","None"])

ws2.append(["Dropdown",119,"","","","Top 3 priority issues","",
            "Water crisis","Air pollution","Drug addiction",
            "Unemployment","Youth migration","Inflation",
            "Healthcare","Education","Corruption","Other","No opinion"])

ws2.append(["Dropdown",120,"","","","First-time benefit under Mann govt","",
            "Education","Electricity","Employment","Healthcare",
            "Transport","Farmer support","Governance","Other","None"])
ws2.append(["Single Select",121,"","","","Received this service 2017-22?","","Yes","No","Not Sure"])
ws2.append(["Single Select",122,"","","","Received this service 2012-17?","","Yes","No","Not Sure"])

ws2.append(["Dropdown",123,"","","","Agricultural land status","",
            "Owns land","Leases land","Works for wages"])
ws2.append(["Dropdown",124,"","","","Total landholding size","",
            "Less than 1 acre","1-5 acres","5-10 acres",
            "More than 10 acres","Can't Say"])
ws2.append(["Dropdown",125,"","","","Crops cultivated","",
            "Wheat","Rice","Maize","Sugarcane",
            "Vegetables","Fruits","Other","Can't Say"])
ws2.append(["Single Select",126,"","","","Want to diversify crops?","","Yes","No","Not Sure"])
ws2.append(["Dropdown",127,"","","","Source of irrigation","",
            "Canal","Tube well","Rain-fed","Other","Can't Say"])
ws2.append(["Dropdown",128,"","","","Water shortage for irrigation","",
            "Frequently","Occasionally","Rarely","No issues","Can't Say"])
ws2.append(["Single Select",129,"","","","Canal irrigation improved since 2022?","","Yes","No","Can't Say"])
ws2.append(["Dropdown",130,"","","","Farm electricity hours per day","",
            "8 hours day","8 hours mixed","Less than 8","More than 8","Can't Say"])
ws2.append(["Single Select",131,"","","","Farm electricity improved since 2022?","","Yes","No","Can't Say"])


# (You can continue appending remaining rows in the same format)

# ---------------- Sheet 3: options ----------------
ws3 = wb.create_sheet("options")
ws3.append(["Party","Place","State"])
ws3.append(["BJP","Uttarakhand","Delhi"])
ws3.append(["INC","Delhi","Punjab"])
ws3.append(["AAP","Himanchal","Gujarat"])
ws3.append(["Others","Kashmir","Goa"])

# ---------------- Save file ----------------
wb.save("Updated_ReadData1.xlsx")

print("Excel file created successfully!")


